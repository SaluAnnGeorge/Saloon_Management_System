from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Avg
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import JsonResponse
from django.template.loader import render_to_string
from .models import Category, Service, Review, Contact
from .forms import ReviewForm, ServiceReviewForm
from django.db import IntegrityError

def index(request):
    services = Service.objects.all()
    return render(request, 'saloon/saloon.html', {'services': services})

def about(request):
    return render(request, 'saloon/about.html')




from django.core.mail import send_mail
from django.conf import settings
from django.contrib import messages
from django.shortcuts import render
from .forms import ContactForm

def contact(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            # Extract data from the form
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            subject = form.cleaned_data['subject']
            message = form.cleaned_data['message']

            # Send email
            send_mail(
                f'Contact Us Form: {subject}',  # Email subject
                f'Message from {name} ({email}):\n\n{message}',  # Email body
                settings.DEFAULT_FROM_EMAIL,  # From email (configured in settings.py)
                ['saluanngeorge@gmail.com'],  # Recipient email
                fail_silently=False,
            )

            # Show success message
            messages.success(request, 'Your message has been sent successfully!')
            return render(request, 'saloon/contact.html', {'form': form})
        else:
            # Show error message
            messages.error(request, 'There was an error with your submission. Please try again.')
    else:
        form = ContactForm()

    return render(request, 'saloon/contact.html', {'form': form})
def gallery(request):
    return render(request, 'saloon/gallery.html')

def error(request):
    return render(request, 'saloon/404.html')

def service(request, category_id=None):
    if category_id:
        category = get_object_or_404(Category, id=category_id)
        services = Service.objects.filter(category=category)
    else:
        services = Service.objects.all()
    
    # Add average rating for each service
    for service in services:
        service.avg_rating = service.get_average_rating()
        service.review_count = service.get_review_count()
    
    categories = Category.objects.all()
    return render(request, 'saloon/service.html', {
        'services': services, 
        'categories': categories
    })

def service_by_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    services = Service.objects.filter(category=category)
    context = {
        'category': category,
        'services': services
    }
    return render(request, 'saloon/service.html', context)

def service_detail(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    reviews = service.reviews.select_related('user').order_by('-created_at')
    
    # Get review statistics
    avg_rating = service.get_average_rating()
    reviews_count = service.get_review_count()
    
    # Check if user has already reviewed
    user_review = None
    if request.user.is_authenticated:
        user_review = reviews.filter(user=request.user).first()
    
    # Handle AJAX requests for loading more reviews
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        page = request.GET.get('page', 1)
        paginator = Paginator(reviews, 6)
        try:
            reviews_page = paginator.page(page)
        except (PageNotAnInteger, EmptyPage):
            return JsonResponse({'reviews': [], 'has_next': False})
        
        reviews_data = []
        for review in reviews_page:
            reviews_data.append({
                'html': render_to_string('saloon/partials/review_card.html', {'review': review})
            })
        
        return JsonResponse({
            'reviews': reviews_data,
            'has_next': reviews_page.has_next()
        })
    
    context = {
        'service': service,
        'reviews': reviews[:6],  # Show first 6 reviews initially
        'avg_rating': avg_rating,
        'reviews_count': reviews_count,
        'user_review': user_review,
    }
    return render(request, 'saloon/service_detail.html', context)

@login_required
def add_review(request):
    if request.method == 'POST':
        form = ReviewForm(request.POST)
        if form.is_valid():
            try:
                review = form.save(commit=False)
                review.user = request.user
                review.save()
                messages.success(request, 'Thank you! Your review has been added successfully.')
                return redirect('saloon:testimonial')
            except IntegrityError:
                messages.error(request, 'You have already reviewed this service. You can edit your existing review instead.')
                return redirect('saloon:testimonial')
    else:
        form = ReviewForm()
    
    return render(request, 'saloon/add_review.html', {'form': form})

@login_required
def add_service_review(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    if request.method == 'POST':
        form = ServiceReviewForm(request.POST)
        if form.is_valid():
            review = form.save(commit=False)
            review.user = request.user
            review.service = service
            review.save()
            messages.success(request, 'Thank you for your review!')
            return redirect('saloon:service_detail', service_id=service_id)
    else:
        form = ServiceReviewForm()
    
    return render(request, 'saloon/add_review.html', {
        'form': form,
        'service': service
    })

@login_required
def delete_review(request, review_id):
    review = get_object_or_404(Review, id=review_id)
    if request.user == review.user or request.user.is_staff:
        review.delete()
        messages.success(request, 'Review deleted successfully!')
    else:
        messages.error(request, 'You do not have permission to delete this review.')
    return redirect('saloon:testimonial')

@login_required
def edit_review(request, review_id):
    review = get_object_or_404(Review, id=review_id)
    if request.user != review.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to edit this review.')
        return redirect('saloon:testimonial')
        
    if request.method == 'POST':
        form = ReviewForm(request.POST, instance=review)
        if form.is_valid():
            form.save()
            messages.success(request, 'Review updated successfully!')
            return redirect('saloon:testimonial')
    else:
        form = ReviewForm(instance=review)
    
    return render(request, 'saloon/edit_review.html', {'form': form, 'review': review})

@login_required
def edit_review_service(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    review = get_object_or_404(Review, service=service, user=request.user)
    
    if request.method == 'POST':
        rating = request.POST.get('rating')
        comment = request.POST.get('comment')
        
        if not rating or not comment:
            messages.error(request, 'Both rating and comment are required.')
            return redirect('saloon:edit_review', service_id=service_id)
        
        review.rating = int(rating)
        review.comment = comment
        review.save()
        
        messages.success(request, 'Your review has been updated!')
        return redirect('saloon:service_detail', service_id=service_id)
    
    context = {
        'service': service,
        'review': review,
        'rating_choices': Review.RATING_CHOICES
    }
    return render(request, 'saloon/edit_review.html', context)

@login_required
def delete_review_service(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    review = get_object_or_404(Review, service=service, user=request.user)
    
    if request.method == 'POST':
        review.delete()
        messages.success(request, 'Your review has been deleted.')
        return redirect('saloon:service_detail', service_id=service_id)
    
    return render(request, 'saloon/delete_review.html', {'service': service, 'review': review})

@login_required
def my_enquiries(request):
    enquiries = Contact.objects.filter(email=request.user.email).order_by('-created_at')
    return render(request, 'saloon/my_enquiries.html', {'enquiries': enquiries})

def price(request):
    return render(request, 'saloon/price.html')

def team(request):
    return render(request, 'saloon/team.html')

def testimonial(request):
    reviews = Review.objects.select_related('user', 'service').order_by('-created_at')
    return render(request, 'saloon/testimonial.html', {'reviews': reviews})

def appointment(request):
    services = Service.objects.all()
    return render(request, 'saloon/appointment.html', {'services': services})


from django.shortcuts import render
from .models import Service

def search_services(request):
    query = request.GET.get('q', '')  # Get the search query from the URL
    if query:
        services = Service.objects.filter(name__icontains=query)  # Filter services based on the search query
    else:
        services = Service.objects.all()  # If no query, show all services
    return render(request, 'saloon/service_search.html', {'services': services, 'query': query})

from django.http import HttpResponse
from django.shortcuts import render
import openpyxl
from openpyxl.utils import get_column_letter
from .models import User
from appointment.models import Appointment
from datetime import datetime

def customer_report(request):
    # Get all users (customers) from the database
    customers = User.objects.all()

    if 'export' in request.GET:
        # Create an Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Report"

        # Write the headers
        headers = ['Name', 'Email', 'Phone', 'Appointment Date', 'Service', 'Appointment Status', 'Payment Status']
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header

        # Write data to the Excel sheet
        for row_num, customer in enumerate(customers, 2):
            appointments = Appointment.objects.filter(customer=customer)
            
            for appointment in appointments:
                ws[f"A{row_num}"] = customer.full_name
                ws[f"B{row_num}"] = customer.email
                ws[f"C{row_num}"] = customer.phone
                ws[f"D{row_num}"] = appointment.appointment_date
                ws[f"E{row_num}"] = appointment.service.name
                ws[f"F{row_num}"] = appointment.status
                ws[f"G{row_num}"] = appointment.payment_status
                row_num += 1

        # Prepare the response to download the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Customer_Report.xlsx'
        wb.save(response)
        return response

    # Context for rendering the HTML page
    context = {
        'customers': customers,
    }
    return render(request, 'saloon/customer_report.html', context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Avg
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import JsonResponse
from django.template.loader import render_to_string
from .models import Category, Service, Review, Contact
from .forms import ReviewForm, ServiceReviewForm
from django.db import IntegrityError

def index(request):
    services = Service.objects.all()
    return render(request, 'saloon/saloon.html', {'services': services})

def about(request):
    return render(request, 'saloon/about.html')

# def contact(request):
#     if request.method == 'POST':
#         print("Received POST request with data:", request.POST)  # Debug print
#         name = request.POST.get('name')
#         email = request.POST.get('email')
#         subject = request.POST.get('subject')
#         message = request.POST.get('message')
        
#         print(f"Processing contact form - Name: {name}, Email: {email}, Subject: {subject}")  # Debug print
        
#         try:
#             contact = Contact.objects.create(
#                 name=name,
#                 email=email,
#                 subject=subject,
#                 message=message
#             )
#             print(f"Contact created with ID: {contact.id}")  # Debug print
#             messages.success(request, 'Your message has been sent successfully! We will get back to you soon.')
#         except Exception as e:
#             print(f"Error creating contact: {str(e)}")  # Debug print
#             messages.error(request, 'There was an error sending your message. Please try again.')
#         return redirect('saloon:contact')
        
#     return render(request, 'saloon/contact.html')


from django.core.mail import send_mail
from django.conf import settings
from django.contrib import messages
from django.shortcuts import render
from .forms import ContactForm

def contact(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            # Extract data from the form
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            subject = form.cleaned_data['subject']
            message = form.cleaned_data['message']

            # Send email
            send_mail(
                f'Contact Us Form: {subject}',  # Email subject
                f'Message from {name} ({email}):\n\n{message}',  # Email body
                settings.DEFAULT_FROM_EMAIL,  # From email (configured in settings.py)
                ['saluanngeorge@gmail.com'],  # Recipient email
                fail_silently=False,
            )

            # Show success message
            messages.success(request, 'Your message has been sent successfully!')
            return render(request, 'saloon/contact.html', {'form': form})
        else:
            # Show error message
            messages.error(request, 'There was an error with your submission. Please try again.')
    else:
        form = ContactForm()

    return render(request, 'saloon/contact.html', {'form': form})
def gallery(request):
    return render(request, 'saloon/gallery.html')

def error(request):
    return render(request, 'saloon/404.html')

def service(request, category_id=None):
    if category_id:
        category = get_object_or_404(Category, id=category_id)
        services = Service.objects.filter(category=category)
    else:
        services = Service.objects.all()
    
    # Add average rating for each service
    for service in services:
        service.avg_rating = service.get_average_rating()
        service.review_count = service.get_review_count()
    
    categories = Category.objects.all()
    return render(request, 'saloon/service.html', {
        'services': services, 
        'categories': categories
    })

def service_by_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    services = Service.objects.filter(category=category)
    context = {
        'category': category,
        'services': services
    }
    return render(request, 'saloon/service.html', context)

def service_detail(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    reviews = service.reviews.select_related('user').order_by('-created_at')
    
    # Get review statistics
    avg_rating = service.get_average_rating()
    reviews_count = service.get_review_count()
    
    # Check if user has already reviewed
    user_review = None
    if request.user.is_authenticated:
        user_review = reviews.filter(user=request.user).first()
    
    # Handle AJAX requests for loading more reviews
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        page = request.GET.get('page', 1)
        paginator = Paginator(reviews, 6)
        try:
            reviews_page = paginator.page(page)
        except (PageNotAnInteger, EmptyPage):
            return JsonResponse({'reviews': [], 'has_next': False})
        
        reviews_data = []
        for review in reviews_page:
            reviews_data.append({
                'html': render_to_string('saloon/partials/review_card.html', {'review': review})
            })
        
        return JsonResponse({
            'reviews': reviews_data,
            'has_next': reviews_page.has_next()
        })
    
    context = {
        'service': service,
        'reviews': reviews[:6],  # Show first 6 reviews initially
        'avg_rating': avg_rating,
        'reviews_count': reviews_count,
        'user_review': user_review,
    }
    return render(request, 'saloon/service_detail.html', context)

@login_required
def add_review(request):
    if request.method == 'POST':
        form = ReviewForm(request.POST)
        if form.is_valid():
            try:
                review = form.save(commit=False)
                review.user = request.user
                review.save()
                messages.success(request, 'Thank you! Your review has been added successfully.')
                return redirect('saloon:testimonial')
            except IntegrityError:
                messages.error(request, 'You have already reviewed this service. You can edit your existing review instead.')
                return redirect('saloon:testimonial')
    else:
        form = ReviewForm()
    
    return render(request, 'saloon/add_review.html', {'form': form})

@login_required
def add_service_review(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    if request.method == 'POST':
        form = ServiceReviewForm(request.POST)
        if form.is_valid():
            review = form.save(commit=False)
            review.user = request.user
            review.service = service
            review.save()
            messages.success(request, 'Thank you for your review!')
            return redirect('saloon:service_detail', service_id=service_id)
    else:
        form = ServiceReviewForm()
    
    return render(request, 'saloon/add_review.html', {
        'form': form,
        'service': service
    })

@login_required
def delete_review(request, review_id):
    review = get_object_or_404(Review, id=review_id)
    if request.user == review.user or request.user.is_staff:
        review.delete()
        messages.success(request, 'Review deleted successfully!')
    else:
        messages.error(request, 'You do not have permission to delete this review.')
    return redirect('saloon:testimonial')

@login_required
def edit_review(request, review_id):
    review = get_object_or_404(Review, id=review_id)
    if request.user != review.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to edit this review.')
        return redirect('saloon:testimonial')
        
    if request.method == 'POST':
        form = ReviewForm(request.POST, instance=review)
        if form.is_valid():
            form.save()
            messages.success(request, 'Review updated successfully!')
            return redirect('saloon:testimonial')
    else:
        form = ReviewForm(instance=review)
    
    return render(request, 'saloon/edit_review.html', {'form': form, 'review': review})

@login_required
def edit_review_service(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    review = get_object_or_404(Review, service=service, user=request.user)
    
    if request.method == 'POST':
        rating = request.POST.get('rating')
        comment = request.POST.get('comment')
        
        if not rating or not comment:
            messages.error(request, 'Both rating and comment are required.')
            return redirect('saloon:edit_review', service_id=service_id)
        
        review.rating = int(rating)
        review.comment = comment
        review.save()
        
        messages.success(request, 'Your review has been updated!')
        return redirect('saloon:service_detail', service_id=service_id)
    
    context = {
        'service': service,
        'review': review,
        'rating_choices': Review.RATING_CHOICES
    }
    return render(request, 'saloon/edit_review.html', context)

@login_required
def delete_review_service(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    review = get_object_or_404(Review, service=service, user=request.user)
    
    if request.method == 'POST':
        review.delete()
        messages.success(request, 'Your review has been deleted.')
        return redirect('saloon:service_detail', service_id=service_id)
    
    return render(request, 'saloon/delete_review.html', {'service': service, 'review': review})

@login_required
def my_enquiries(request):
    enquiries = Contact.objects.filter(email=request.user.email).order_by('-created_at')
    return render(request, 'saloon/my_enquiries.html', {'enquiries': enquiries})

def price(request):
    return render(request, 'saloon/price.html')

def team(request):
    return render(request, 'saloon/team.html')

def testimonial(request):
    reviews = Review.objects.select_related('user', 'service').order_by('-created_at')
    return render(request, 'saloon/testimonial.html', {'reviews': reviews})

def appointment(request):
    services = Service.objects.all()
    return render(request, 'saloon/appointment.html', {'services': services})


from django.shortcuts import render
from .models import Service

def search_services(request):
    query = request.GET.get('q', '')  # Get the search query 
    if query:
        services = Service.objects.filter(name__icontains=query)  # Filter services based on the search query
    else:
        services = Service.objects.all()  # If no query, show all services
    return render(request, 'saloon/service_search.html', {'services': services, 'query': query})

from django.http import HttpResponse
from django.shortcuts import render
import openpyxl
from openpyxl.utils import get_column_letter
from .models import User
from appointment.models import Appointment
from datetime import datetime

def customer_report(request):
    # Get all users (customers) from the database
    customers = User.objects.all()

    if 'export' in request.GET:
        # Create an Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Report"

        # Write the headers
        headers = ['Name', 'Email', 'Phone', 'Appointment Date', 'Service', 'Appointment Status', 'Payment Status']
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header

        # Write data to the Excel sheet
        for row_num, customer in enumerate(customers, 2):
            appointments = Appointment.objects.filter(customer=customer)
            
            for appointment in appointments:
                ws[f"A{row_num}"] = customer.full_name
                ws[f"B{row_num}"] = customer.email
                ws[f"C{row_num}"] = customer.phone
                ws[f"D{row_num}"] = appointment.appointment_date
                ws[f"E{row_num}"] = appointment.service.name
                ws[f"F{row_num}"] = appointment.status
                ws[f"G{row_num}"] = appointment.payment_status
                row_num += 1

        # Prepare the response to download the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Customer_Report.xlsx'
        wb.save(response)
        return response

    # Context for rendering the HTML page
    context = {
        'customers': customers,
    }
    return render(request, 'saloon/customer_report.html', context)


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
import google.generativeai as genai
from .models import Service, Category

genai.configure(api_key="AIzaSyCmlRAihQuprZlWG_P5_jz33__5_ZGYK_M")

def personalized_recommendation(request):
    recommendation = None
    
    if request.method == "POST":
        service_type = request.POST.get('service_type')
        hair_type = request.POST.get('hair_type')
        skin_type = request.POST.get('skin_type')
        budget = request.POST.get('budget')
        
        user_input = f"""
        Service Type: {service_type}
        Hair Type: {hair_type}
        Skin Type: {skin_type}
        Budget: {budget}
        Recommend a salon service based on the above criteria.
        """
        
        model = genai.GenerativeModel(model_name="gemini-1.5-flash")
        chat_session = model.start_chat(history=[])
        response = chat_session.send_message(user_input)
        
        recommendation = response.text
    
    return render(request, 'saloon/service_recommendation.html', {'recommendation': recommendation})






from django.db.models import Q
from django.shortcuts import render, redirect
from .models import Service

def search_service(request):
    if request.method == "POST":
        query = request.POST.get("search_query", "").strip().lower()  # Get search text

        # Filter services based on search_query
        services = Service.objects.filter(
            name__icontains=query  # Search for service name
        )

        return render(request, "saloon/search_results.html", {'services': services, 'query': query})

    return redirect("saloon:index")




import os
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
from django.shortcuts import render
from django.db.models import Count
from django.conf import settings
from appointment.models import Appointment

def generate_appointment_charts(request):
    # Get all appointment data
    appointments = Appointment.objects.all()

    if not appointments.exists():
        return render(request, "appointment/appointment_charts.html", {"error": "No appointment data available."})

    # Convert QuerySet to DataFrame
    df = pd.DataFrame(list(appointments.values("appointment_date", "service__name")))
    
    if df.empty:
        return render(request, "appointment/appointment_charts.html", {"error": "No appointment data available."})

    df["appointment_date"] = pd.to_datetime(df["appointment_date"], errors="coerce")
    df = df.dropna(subset=["appointment_date"])  # Remove any NaT values

    # Ensure the directory exists for saving charts
    chart_dir = os.path.join(settings.MEDIA_ROOT, "charts")
    os.makedirs(chart_dir, exist_ok=True)

    ### 1. **Bar Chart: Number of Appointments Per Day**
    daily_counts = df.groupby(df["appointment_date"].dt.date).size()
    plt.figure(figsize=(10, 5))
    daily_counts.plot(kind="bar", color="blue", alpha=0.7)
    plt.title("Number of Appointments Per Day")
    plt.xlabel("Date")
    plt.ylabel("Total Appointments")
    plt.xticks(rotation=45)
    plt.tight_layout()
    daily_chart_path = os.path.join(chart_dir, "daily_appointments.png")
    plt.savefig(daily_chart_path)
    plt.clf()

    ### 2. **Pie Chart: Distribution of Services Booked**
    service_counts = df["service__name"].value_counts()
    plt.figure(figsize=(8, 8))
    plt.pie(service_counts, labels=service_counts.index, autopct="%1.1f%%", startangle=140, colors=sns.color_palette("pastel"))
    plt.title("Distribution of Services Booked")
    plt.axis("equal")
    pie_chart_path = os.path.join(chart_dir, "service_distribution.png")
    plt.savefig(pie_chart_path)
    plt.clf()


    # Convert file paths to URLs for displaying images in the template
    chart_urls = {
        "daily_chart_url": settings.MEDIA_URL + "charts/daily_appointments.png",
        "pie_chart_url": settings.MEDIA_URL + "charts/service_distribution.png",
       
    }

    return render(request, "saloon/appointment_charts.html", chart_urls)


